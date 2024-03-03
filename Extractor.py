import openpyxl
import os, datetime, time
import docx
from shutil import copy2

result_column_id = ['C', 'D', ]

subtotal_spreadsheet_path = os.path.join( os.getcwd(), "ExtractResult.xlsx")
if os.path.exists(subtotal_spreadsheet_path):
    try:
        os.remove(subtotal_spreadsheet_path)
    except Exception as ex:
        print("can not remove the result file, please check if it is opening.")
        exit()
print(subtotal_spreadsheet_path)


class ProjInfo:
    #proj_code = ""
    #end_date = None
    def __init__(self, code, end_date):
        self.proj_code = code
        # the date of the end of the proj
        self.end_date = end_date
        self.Column_C = [0] * 37
        self.Column_D = [0] * 37
        self.Column_E = [0] * 37
        self.Column_F = [0] * 37
        self.Column_G = [0] * 37
        self.Column_H = [0] * 37
        self.Column_I = [0] * 37
        self.vender_list = []

proj_info_dict = {}

spreadsheet_list = []
report_list = []
for (root, dirs, files) in os.walk(os.getcwd()):
    for file in files:
        if file.startswith("~$"):
            continue
        if file.endswith(".xlsx"):
            spreadsheet_list.append(os.path.join(root, file))
        if file.endswith("设计批复文件.docx"):
            report_list.append(os.path.join(root, file))


for report in report_list:
    doc = docx.Document(report)
    tb = doc.tables[0]

    proj_code = ""
    proj_end_date = ""
    for r in tb.rows:
        if r.cells[0].text == "项目编码":
            proj_code = r.cells[1].text.upper()
        if r.cells[1].text == "交付使用时间":
            proj_end_date = datetime.datetime.strptime(r.cells[3].text, "%Y-%m-%d")
            
    curr_info = ProjInfo(proj_code, proj_end_date)
    proj_info_dict[proj_code] = curr_info


for spreadsheet in spreadsheet_list:
    proj_code = os.path.basename(spreadsheet).split('.')[0].split('_')[0].upper()
    if not proj_code in proj_info_dict.keys():
        proj_info_dict[proj_code] = ProjInfo(proj_code, None)
    curr_proj_info = proj_info_dict[proj_code]
    
    workbook = openpyxl.load_workbook(spreadsheet)
    worksheet = workbook["工程项目竣工决算总表"]
    for r in range(0, 37):
        curr_proj_info.Column_C[r] = (float(worksheet["C{}".format(r+2)].value))
        curr_proj_info.Column_D[r] = (float(worksheet["D{}".format(r+2)].value))
        curr_proj_info.Column_E[r] = (float(worksheet["E{}".format(r+2)].value))
        curr_proj_info.Column_F[r] = (float(worksheet["F{}".format(r+2)].value))
        curr_proj_info.Column_G[r] = (float(worksheet["G{}".format(r+2)].value))
        curr_proj_info.Column_H[r] = (float(worksheet["H{}".format(r+2)].value))
        curr_proj_info.Column_I[r] = (float(worksheet["I{}".format(r+2)].value))

    # get vender list:
    for sheet_name in ["成本单(费用清单)", "成本单(设备清单)", "成本单"]:
        sheet = workbook[sheet_name]
        for c in sheet["C"]:
            curr_proj_info.vender_list.append(c.value)
    curr_proj_info.vender_list = list(set(curr_proj_info.vender_list))
    curr_proj_info.vender_list = [i for i in curr_proj_info.vender_list if i]
    curr_proj_info.vender_list = [i for i in curr_proj_info.vender_list if i != "供应商"]

subtotal_info = ProjInfo("SubTotal", None)

for proj_info in proj_info_dict.values():
    for i in range(0, 37):
        subtotal_info.Column_C[i] += proj_info.Column_C[i]
        subtotal_info.Column_D[i] += proj_info.Column_D[i]
        subtotal_info.Column_E[i] += proj_info.Column_E[i]
        subtotal_info.Column_F[i] += proj_info.Column_F[i]
        subtotal_info.Column_G[i] += proj_info.Column_G[i]
        subtotal_info.Column_H[i] += proj_info.Column_H[i]
        subtotal_info.Column_I[i] += proj_info.Column_I[i]

#print(subtotal_info.proj_code, "(", subtotal_info.end_date, ")\n\t", subtotal_info.Column_C, "\n\t", subtotal_info.Column_D, "\n\t", subtotal_info.Column_E, "\n\t", subtotal_info.Column_F)

# rounded = round(8.88888, 2) 
# print(rounded) 
  
# formatted = "{:.3f}".format(9.999) 
# print(formatted)

# new extract result spreadsheet
fp = copy2(os.path.join(os.getcwd(), "template.binary"), subtotal_spreadsheet_path)
time.sleep(5)

if not os.path.exists(subtotal_spreadsheet_path):
    print("extract result file not exists.")
    exit()

workbook = openpyxl.load_workbook(subtotal_spreadsheet_path)
sheet = workbook.active
if sheet is None:
    print("error")
    exit()

# 工程项目竣工决算总表 
for r in range(0, 37):
    sheet["C{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_C[r])
    sheet["D{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_D[r])
    sheet["E{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_E[r])
    sheet["F{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_F[r])
    sheet["G{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_G[r])
    sheet["H{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_H[r])
    sheet["I{}".format(r+2)] = "{:.2f}".format(subtotal_info.Column_I[r])

# 批复概算 
sheet = workbook["批复概算"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_C[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_C[r])

# 审计调整概算 
sheet = workbook["审计调整概算"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_D[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_D[r])

# 实际执行概算 
sheet = workbook["实际执行概算"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_E[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_E[r])

# 决算送审金额 
sheet = workbook["决算送审金额"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_F[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_F[r])

# 结算审计未调整金额 
sheet = workbook["结算审计未调整金额"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_G[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_G[r])

# 决算审计调整金额 
sheet = workbook["决算审计调整金额"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_H[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_H[r])

# 决算审定金额
sheet = workbook["决算审定金额"]
for r in range(0, 37):
    curr_column = 3
    sheet.cell(row = 1, column=curr_column, value = "合计")
    sheet.cell(row = r+2, column=curr_column, value = subtotal_info.Column_I[r])
    for v in proj_info_dict.values():
        curr_column += 1
        sheet.cell(row = 1, column=curr_column, value = v.proj_code)
        sheet.cell(row = r+2, column=curr_column, value = v.Column_I[r])

# 供应商
sheet = workbook.create_sheet( title= "供应商")
r = 0
for info in proj_info_dict.values():
    c = 1
    r += 1
    sheet.cell(row = r, column=c, value = info.proj_code)
    c += 1
    sheet.cell(row=r, column=c, value = info.end_date)
    for vender in info.vender_list:
        c += 1
        sheet.cell(row = r, column=c, value = vender)
        

workbook.save(subtotal_spreadsheet_path)
workbook.close()